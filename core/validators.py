from openpyxl import load_workbook

def _norm_col(c):
    return (
        str(c or "")
        .strip()
        .lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )

def _headers(ws):
    return {
        _norm_col(cell.value): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value is not None and str(cell.value).strip() != ""
    }

def _cell_text(value):
    return str(value or "").strip()

def validar_maestro(maestro_path):
    errores = []
    advertencias = []

    try:
        wb = load_workbook(maestro_path, read_only=True, data_only=True)
    except Exception as e:
        return False, [f"No pude abrir el Excel del maestro. Error: {e}"], []

    hojas_requeridas = ["Productos", "Aliases", "Exclusiones", "Configuración"]
    hojas = set(wb.sheetnames)

    for hoja in hojas_requeridas:
        if hoja not in hojas:
            errores.append(f"Falta la hoja obligatoria: {hoja}")

    if errores:
        wb.close()
        return False, errores, advertencias

    # PRODUCTOS
    ws = wb["Productos"]
    cols = _headers(ws)

    requeridas_productos = [
        "producto base",
        "codigo compra",
        "producto compra",
        "compra minima",
    ]

    for req in requeridas_productos:
        if req not in cols:
            errores.append(f"Hoja Productos: falta columna '{req}'.")

    if not errores:
        col_producto = cols["producto base"]
        col_codigo = cols["codigo compra"]
        col_prod_compra = cols["producto compra"]
        col_minima = cols["compra minima"]

        filas_validas = 0
        codigos_validos = 0
        compras_minimas_validas = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            producto = row[col_producto - 1] if len(row) >= col_producto else None
            codigo = row[col_codigo - 1] if len(row) >= col_codigo else None
            prod_compra = row[col_prod_compra - 1] if len(row) >= col_prod_compra else None
            compra_min = row[col_minima - 1] if len(row) >= col_minima else None

            if _cell_text(producto) or _cell_text(codigo) or _cell_text(prod_compra):
                filas_validas += 1

            if _cell_text(codigo):
                codigos_validos += 1

            try:
                if compra_min is not None and float(compra_min) > 0:
                    compras_minimas_validas += 1
            except Exception:
                pass

        if filas_validas < 5:
            errores.append("Hoja Productos: hay muy pocas filas válidas. No parece ser el Maestro real.")

        if codigos_validos < 5:
            errores.append("Hoja Productos: hay muy pocos códigos de compra válidos.")

        if compras_minimas_validas < 5:
            errores.append("Hoja Productos: hay muy pocas compras mínimas numéricas válidas.")

    # ALIASES
    ws = wb["Aliases"]
    cols = _headers(ws)

    if not ("alias detectado" in cols or "alias" in cols or "nombre original" in cols):
        errores.append("Hoja Aliases: falta columna de alias.")

    if "producto base" not in cols:
        errores.append("Hoja Aliases: falta columna 'Producto Base'.")

    # CONFIGURACION
    ws = wb["Configuración"]
    cols = _headers(ws)

    if "parametro" not in cols or "valor" not in cols:
        errores.append("Hoja Configuración: debe tener columnas Parámetro y Valor.")
    else:
        col_param = cols["parametro"]
        parametros = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[col_param - 1] if len(row) >= col_param else None
            if _cell_text(val):
                parametros.add(_cell_text(val).lower())

        esperados = {
            "semanas objetivo",
            "tiempo de reposición",
            "tiempo de reposicion",
            "días analizados",
            "dias analizados",
        }

        if len(parametros.intersection(esperados)) < 2:
            errores.append("Hoja Configuración: no encontré parámetros esperados del Maestro.")

    wb.close()
    return len(errores) == 0, errores, advertencias


def _parse_price_ar(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None

    s = str(value).strip()
    if not s:
        return None

    s = s.replace("$", "").replace(" ", "")

    try:
        if "," in s:
            s2 = s.replace(".", "").replace(",", ".")
            return float(s2)
        return float(s)
    except Exception:
        return None


def _codigo_ok(value):
    if value is None:
        return False

    s = str(value).strip()
    if not s:
        return False

    try:
        n = int(float(s))
        return n > 0
    except Exception:
        return False


def validar_carrito(carrito_path):
    errores = []
    advertencias = []

    try:
        wb = load_workbook(carrito_path, read_only=True, data_only=True)
    except Exception as e:
        return False, [f"No pude abrir el Excel del carrito. Error: {e}"], []

    if len(wb.sheetnames) < 1:
        wb.close()
        return False, ["El carrito debe tener al menos una hoja."], []

    ws = wb[wb.sheetnames[0]]

    if ws.max_column < 10:
        wb.close()
        return False, ["El carrito debe tener al menos 10 columnas. Se espera la estructura del Modelo de Carrito."], []

    # Validación estructural del Modelo de Carrito real.
    # El motor usa B=código e I=precio, pero para evitar falsos positivos validamos encabezados.
    headers = [
        _norm_col(ws.cell(row=1, column=i).value)
        for i in range(1, 11)
    ]

    expected = {
        0: ["grupoproducto", "grupo producto"],
        1: ["codigo", "código"],
        2: ["cantidad"],
        3: ["deposito", "depósito"],
        4: ["descripcion", "descripción"],
        5: ["cubicaje"],
        6: ["peso"],
        7: ["impuestoiva", "impuesto iva", "iva"],
        8: ["precio"],
        9: ["disponible"],
    }

    for idx, posibles in expected.items():
        if headers[idx] not in [_norm_col(x) for x in posibles]:
            errores.append(
                f"Encabezado inválido en columna {idx + 1}. "
                f"Esperado: {posibles[0]}. Detectado: '{headers[idx]}'."
            )

    if errores:
        wb.close()
        return False, errores, advertencias

    codigos_validos = 0
    precios_validos = 0
    descripciones_validas = 0
    pares_validos = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = row[1] if len(row) >= 2 else None
        descripcion = row[4] if len(row) >= 5 else None
        precio = row[8] if len(row) >= 9 else None

        codigo_valido = _codigo_ok(codigo)

        descripcion_valida = (
            descripcion is not None
            and str(descripcion).strip() != ""
            and len(str(descripcion).strip()) >= 3
        )

        precio_parseado = _parse_price_ar(precio)
        precio_valido = precio_parseado is not None and precio_parseado > 0

        if codigo_valido:
            codigos_validos += 1

        if descripcion_valida:
            descripciones_validas += 1

        if precio_valido:
            precios_validos += 1

        if codigo_valido and descripcion_valida and precio_valido:
            pares_validos += 1

    if codigos_validos == 0:
        errores.append("No encontré códigos válidos en la columna B.")

    if descripciones_validas == 0:
        errores.append("No encontré descripciones válidas en la columna E.")

    if precios_validos == 0:
        errores.append("No encontré precios numéricos válidos en la columna I.")

    if pares_validos == 0:
        errores.append("No encontré ninguna fila válida con código, descripción y precio.")

    if pares_validos < 10:
        errores.append(
            f"Encontré solo {pares_validos} filas válidas. "
            "No parece ser el Modelo de Carrito completo."
        )

    if pares_validos > 0:
        advertencias.append(f"Filas válidas con código, descripción y precio: {pares_validos}.")
        advertencias.append(f"Códigos detectados en columna B: {codigos_validos}.")
        advertencias.append(f"Precios detectados en columna I: {precios_validos}.")

    wb.close()
    return len(errores) == 0, errores, advertencias
