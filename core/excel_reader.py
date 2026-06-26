from io import BytesIO
from openpyxl import load_workbook

def _has_value(row):
    return any(cell is not None and str(cell).strip() != "" for cell in row)

def summarize_workbook(content_bytes: bytes):
    wb = load_workbook(BytesIO(content_bytes), read_only=True, data_only=True)
    summary = {
        "sheet_count": len(wb.sheetnames),
        "sheets": [],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_with_data = 0

        for row in ws.iter_rows(values_only=True):
            if _has_value(row):
                rows_with_data += 1

        summary["sheets"].append({
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "rows_with_data": rows_with_data,
        })

    wb.close()
    return summary

def summarize_carrito(content_bytes: bytes):
    wb = load_workbook(BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    filas_con_datos = 0
    codigos_validos = 0
    precios_validos = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if _has_value(row):
            filas_con_datos += 1

        codigo = row[1] if len(row) >= 2 else None
        precio = row[8] if len(row) >= 9 else None

        if codigo is not None and str(codigo).strip() != "":
            codigos_validos += 1

        try:
            if precio is not None and str(precio).strip() != "":
                float(precio)
                precios_validos += 1
        except Exception:
            pass

    summary = {
        "sheet": wb.sheetnames[0],
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "rows_with_data": filas_con_datos,
        "valid_codes_col_b": codigos_validos,
        "valid_prices_col_i": precios_validos,
    }

    wb.close()
    return summary
