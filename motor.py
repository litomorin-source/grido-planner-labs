
from pathlib import Path
import math
import re
import unicodedata
import tempfile
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Archivo opcional con precios: columna B = código pack, columna I = precio.
DEFAULT_CARRITO_FILE = Path(__file__).resolve().parent / "datos" / "Modelo_de_Carrito.xlsx"


def load_precios_carrito(carrito_path=DEFAULT_CARRITO_FILE):
    """
    Lee el archivo Modelo de Carrito.
    Columna B: código del pack.
    Columna I: precio del pack.
    Devuelve dict {codigo: precio}.
    """
    if not Path(carrito_path).exists():
        return {}

    try:
        df = pd.read_excel(carrito_path)
        if df.shape[1] < 9:
            return {}

        codigo_col = df.columns[1]  # B
        precio_col = df.columns[8]  # I

        precios = {}
        for _, row in df.iterrows():
            codigo = str(row[codigo_col]).strip()
            if codigo.lower() in {"nan", ""}:
                continue

            # normalizar códigos tipo 4000147.0 -> 4000147
            try:
                codigo_norm = str(int(float(codigo)))
            except Exception:
                codigo_norm = codigo

            precio_raw = row[precio_col]

            # Parser robusto para precios argentinos:
            # - Si Excel ya lo leyó como número, se usa directo.
            # - Si viene como texto "42.000,50", se interpreta como 42000.50.
            # - Si viene como texto "42000.50", se interpreta como 42000.50.
            if isinstance(precio_raw, (int, float)) and pd.notna(precio_raw):
                precio = float(precio_raw)
            else:
                s = str(precio_raw).strip()
                s = s.replace("$", "").replace(" ", "")

                if "," in s:
                    # Formato argentino: 42.000,50
                    s = s.replace(".", "").replace(",", ".")
                else:
                    # Sin coma decimal: dejar punto como decimal si existe.
                    # Si el punto era separador de miles, Excel normalmente ya lo hubiera leído como texto con coma decimal.
                    pass

                precio = pd.to_numeric(s, errors="coerce")

            if pd.notna(precio):
                precios[codigo_norm] = int(float(precio))

        return precios

    except Exception:
        return {}


def normalizar_codigo_compra(codigo):
    try:
        return str(int(float(codigo)))
    except Exception:
        return str(codigo).strip()



def strip_accents(text):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(text))
        if unicodedata.category(c) != "Mn"
    )


def normalize(text):
    s = strip_accents(str(text)).lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def truthy(value):
    return normalize(value) in {"si", "sí", "s", "yes", "true", "1", "activo"}


def to_num_ar(series):
    return pd.to_numeric(
        series.astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0)


def safe_num(value, default=0):
    try:
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def ceil_to_int(value):
    if value <= 0 or pd.isna(value):
        return 0
    return int(math.ceil(value))


def find_column(df, candidates):
    normalized = {normalize(c): c for c in df.columns}
    for candidate in candidates:
        cand = normalize(candidate)
        if cand in normalized:
            return normalized[cand]
    for original in df.columns:
        no = normalize(original)
        if any(normalize(c) in no for c in candidates):
            return original
    return None


def load_config(maestro_path, overrides=None):
    try:
        cfg = pd.read_excel(maestro_path, sheet_name="Configuración")
    except Exception:
        cfg = pd.DataFrame()

    defaults = {
        "Semanas objetivo": 4,
        "Tiempo de reposición": 1,
        "Días analizados": 14,
        "Valor sin rotación": 100,
        "Incluir congelados": "Sí",
        "Mostrar sin clasificar": "Sí",
    }

    if not cfg.empty:
        param_col = find_column(cfg, ["Parámetro", "Parametro"])
        value_col = find_column(cfg, ["Valor"])
        if param_col and value_col:
            for _, row in cfg.iterrows():
                p = str(row[param_col]).strip()
                if p and p.lower() != "nan":
                    defaults[p] = row[value_col]

    if overrides:
        defaults.update(overrides)

    return defaults


def load_productos(maestro_path):
    productos = pd.read_excel(maestro_path, sheet_name="Productos")
    productos.columns = [str(c).strip() for c in productos.columns]

    col_producto = find_column(productos, ["Producto Base", "Producto"])
    col_codigo = find_column(productos, ["Código Compra", "Codigo Compra", "codigo compra"])
    col_producto_compra = find_column(productos, ["Producto Compra", "prodcuto compra", "Descripción Compra", "Descripcion Compra"])
    col_minima = find_column(productos, ["Compra Mínima", "Compra Minima", "compra minima"])
    col_tipo = find_column(productos, ["Tipo Producto", "Tipo"])
    col_activo = find_column(productos, ["Activo"])
    col_excluir = find_column(productos, ["Excluir"])

    required = [col_producto, col_codigo, col_producto_compra, col_minima]
    if any(c is None for c in required):
        raise ValueError("La hoja Productos del maestro debe tener: Producto Base, Código Compra, Producto Compra y Compra Mínima.")

    out = pd.DataFrame()
    out["Producto Base"] = productos[col_producto].astype(str).apply(normalize)
    out["Código Compra"] = productos[col_codigo]
    out["Producto Compra"] = productos[col_producto_compra]
    out["Compra Mínima"] = pd.to_numeric(productos[col_minima], errors="coerce").fillna(1)
    out["Tipo Producto"] = productos[col_tipo] if col_tipo else ""
    out["Activo"] = productos[col_activo].apply(truthy) if col_activo else True
    out["Excluir"] = productos[col_excluir].apply(truthy) if col_excluir else False
    out = out[out["Producto Base"].notna() & (out["Producto Base"] != "")]
    return out


def load_aliases(maestro_path):
    try:
        aliases = pd.read_excel(maestro_path, sheet_name="Aliases")
    except Exception:
        return pd.DataFrame(columns=["Alias Detectado", "Alias Normalizado", "Producto Base", "Tipo Presentación", "Pack Detectado", "Regla Conversión"])

    aliases.columns = [str(c).strip() for c in aliases.columns]
    col_alias = find_column(aliases, ["Alias Detectado", "Alias", "Nombre original"])
    col_producto = find_column(aliases, ["Producto Base"])
    col_tipo = find_column(aliases, ["Tipo Presentación", "Tipo presentacion"])
    col_pack = find_column(aliases, ["Pack Detectado"])
    col_regla = find_column(aliases, ["Regla Conversión", "Regla conversion"])

    if col_alias is None or col_producto is None:
        return pd.DataFrame(columns=["Alias Detectado", "Alias Normalizado", "Producto Base", "Tipo Presentación", "Pack Detectado", "Regla Conversión"])

    out = pd.DataFrame()
    out["Alias Detectado"] = aliases[col_alias].astype(str)
    out["Alias Normalizado"] = aliases[col_alias].astype(str).apply(normalize)
    out["Producto Base"] = aliases[col_producto].astype(str).apply(normalize)
    out["Tipo Presentación"] = aliases[col_tipo] if col_tipo else ""
    out["Pack Detectado"] = pd.to_numeric(aliases[col_pack], errors="coerce") if col_pack else np.nan
    out["Regla Conversión"] = aliases[col_regla] if col_regla else ""
    out = out[out["Alias Normalizado"].notna() & (out["Alias Normalizado"] != "")]
    return out


def load_exclusiones(maestro_path):
    try:
        excl = pd.read_excel(maestro_path, sheet_name="Exclusiones")
    except Exception:
        return set()

    col = find_column(excl, ["Producto", "Producto/Opción", "Producto Opcion", "Alias", "Nombre"])
    if col is None:
        col = excl.columns[0]
    return set(excl[col].dropna().astype(str).apply(normalize))


def detect_pack(raw_name):
    match = re.search(r"x\s*(\d+)", normalize(raw_name))
    return int(match.group(1)) if match else None


def is_unit_name(raw_name):
    return bool(re.search(r"x\s*(unidad|un|u)\b", normalize(raw_name)))


def fallback_product_base(raw_name):
    n = normalize(raw_name)
    patterns = [
        r"en caja\s*x\s*\d+\s*(un|u)\b",
        r"caja\s*x\s*\d+\s*(un|u)\b",
        r"x\s*\d+\s*(un|u)\b",
        r"x\s*unidad\b",
        r"x\s*un\b",
        r"x\s*u\b",
        r"x\s*\d+\b",
        r"7\s*800\s*kg",
        r"\bkg\b",
        r"\bgrido\b",
    ]
    for p in patterns:
        n = re.sub(p, "", n)
    return re.sub(r"\s+", " ", n).strip()


def alias_lookup(raw_name, alias_map):
    raw_norm = normalize(raw_name)
    if raw_norm in alias_map:
        return alias_map[raw_norm]
    base = fallback_product_base(raw_name)
    return {"Producto Base": base, "Tipo Presentación": "", "Pack Detectado": np.nan, "Regla Conversión": "", "Alias Encontrado": False}


def infer_conversion(raw_name, product_base, tipo_presentacion, pack_detectado, productos_map):
    raw_norm = normalize(raw_name)
    tipo = normalize(tipo_presentacion)
    tipo_producto = normalize(productos_map.get(product_base, {}).get("Tipo Producto", ""))

    if "palito" in tipo_producto or "palito" in raw_norm:
        p = safe_num(pack_detectado, 0) or detect_pack(raw_name)
        if is_unit_name(raw_name) or "unidad" in tipo:
            return 10, "cantidad / 10"
        if p == 20 or "x20" in tipo:
            return 0.5, "cantidad * 2"
        return 1, "cantidad"

    if "7 800" in raw_norm or "sabor" in tipo_producto:
        return 1, "cantidad"

    if "cerrado" in tipo_producto:
        return 1, "cantidad"

    if is_unit_name(raw_name) or "unidad" in tipo:
        pack = safe_num(pack_detectado, 0)
        if pack <= 0:
            pack = safe_num(productos_map.get(product_base, {}).get("Compra Mínima", 1), 1)
        return pack, f"cantidad / {pack:g}"

    return 1, "cantidad"


def convert_quantity(raw_name, product_base, quantity, tipo_presentacion, pack_detectado, productos_map):
    divisor, regla = infer_conversion(raw_name, product_base, tipo_presentacion, pack_detectado, productos_map)
    if divisor == 0.5:
        return quantity * 2, "cantidad * 2"
    if divisor and divisor != 1:
        return quantity / divisor, regla
    return quantity, regla


def validar_stock(file_path):
    try:
        df = pd.read_csv(file_path, sep=";", encoding="utf-8-sig", nrows=5)
    except Exception as e:
        return False, f"No pude leer el archivo de stock como CSV separado por punto y coma. Error: {e}"
    required = {"Grupo", "Rubro", "SubRubro", "Item", "Stock"}
    missing = [c for c in required if c not in df.columns]
    if missing:
        return False, f"El archivo de stock no parece correcto. Faltan columnas: {', '.join(missing)}"
    return True, "Stock OK"


def validar_sabores(file_path):
    try:
        df = pd.read_excel(file_path, nrows=8)
    except Exception as e:
        return False, f"No pude leer ventas de sabores. Error: {e}"
    if df.shape[1] < 2:
        return False, "El archivo de sabores no parece correcto: tiene menos de 2 columnas."
    return True, "Sabores OK"


def validar_data(file_path):
    try:
        df = pd.read_excel(file_path, nrows=3)
    except Exception as e:
        return False, f"No pude leer data Power BI. Error: {e}"
    text = " ".join([str(x) for x in df.iloc[0].tolist()])
    for needed in ["Categoría", "SubCategoría", "Grupo", "Producto"]:
        if normalize(needed) not in normalize(text):
            return False, "El archivo data no parece ser la exportación Power BI esperada."
    return True, "Data OK"


def read_stock(stock_file, config, alias_map, productos_map, exclusiones):
    stock = pd.read_csv(stock_file, sep=";", encoding="utf-8-sig")
    for col in ["Stock", "Tránsito"]:
        stock[col] = to_num_ar(stock[col]) if col in stock.columns else 0

    incluir_congelados = truthy(config.get("Incluir congelados", "Sí"))
    mask = ((stock["Rubro"] == "Heladería") & (stock["SubRubro"].isin(["Sabores", "Impulsivos"])))
    if incluir_congelados:
        mask = mask | ((stock["Rubro"] == "Congelados") & (stock["Grupo"].isin(["Congelados", "Congelados Multimarca", "Frizzio"])))

    stock = stock[mask].copy()
    if "Grupo" in stock.columns:
        stock = stock[stock["Grupo"] != "Gridos M. Prima."]

    stock = stock[~stock["Item"].astype(str).str.lower().str.contains("desactivado", na=False)].copy()
    stock = stock[~stock["Item"].astype(str).apply(lambda x: normalize(x) in exclusiones or fallback_product_base(x) in exclusiones)].copy()
    stock["Cantidad Stock"] = stock["Stock"] + stock["Tránsito"]

    rows = []
    for _, r in stock.iterrows():
        raw = str(r["Item"])
        info = alias_lookup(raw, alias_map)
        base = info["Producto Base"]
        if base in exclusiones:
            continue
        qty, regla = convert_quantity(raw, base, r["Cantidad Stock"], info.get("Tipo Presentación", ""), info.get("Pack Detectado", np.nan), productos_map)
        rows.append({
            "Origen": "Stock", "Grupo": r.get("Grupo", ""), "SubRubro": r.get("SubRubro", r.get("Rubro", "")),
            "Producto Base": base, "Nombre original": raw, "Cantidad Eq": qty,
            "Regla Conversión": regla, "Alias Encontrado": info.get("Alias Encontrado", False)
        })
    return pd.DataFrame(rows)


def read_sabores(sabores_file, alias_map, productos_map, exclusiones):
    sab = pd.read_excel(sabores_file)
    sab.columns = ["Producto", "Ventas"]
    sab = sab.iloc[2:].copy()
    sab = sab[sab["Producto"].notna()].copy()

    rows = []
    for _, r in sab.iterrows():
        raw = str(r["Producto"])
        base_candidate = fallback_product_base(raw)
        if normalize(raw) in exclusiones or base_candidate in exclusiones:
            continue
        info = alias_lookup(raw, alias_map)
        base = info["Producto Base"]
        if base in exclusiones:
            continue
        qty_raw = safe_num(r["Ventas"], 0)
        qty, regla = convert_quantity(raw, base, qty_raw, info.get("Tipo Presentación", "Sabor 7,800 kg"), info.get("Pack Detectado", np.nan), productos_map)
        rows.append({
            "Origen": "Ventas sabores", "Grupo": "Helado", "SubRubro": "Sabores",
            "Producto Base": base, "Nombre original": raw, "Cantidad Eq": qty,
            "Regla Conversión": regla, "Alias Encontrado": info.get("Alias Encontrado", False)
        })
    return pd.DataFrame(rows)


def read_data(data_file, config, alias_map, productos_map, exclusiones):
    data = pd.read_excel(data_file)
    data.columns = [
        "Categoria","SubCategoria","Grupo","Producto",
        "PL_Cant","PL_Fact","PL_Kilos","PL_Porc",
        "Promo_Cant","Promo_Fact","Promo_Kilos","Promo_Porc",
        "Total_Cantidad","Total_Fact","Total_Kilos","Total_Porc"
    ]
    data = data.iloc[1:].copy()
    for c in ["Categoria", "SubCategoria", "Grupo"]:
        data[c] = data[c].ffill()

    incluir_congelados = truthy(config.get("Incluir congelados", "Sí"))
    mask = ((data["Categoria"] == "Heladería") & (data["SubCategoria"] == "Impulsivos"))
    if incluir_congelados:
        mask = mask | ((data["Categoria"] == "Congelados") & (data["Grupo"].isin(["Congelados Multimarca", "Frizzio"])))

    data = data[mask & data["Producto"].notna() & (~data["Producto"].astype(str).str.lower().eq("total"))].copy()

    rows = []
    for _, r in data.iterrows():
        raw = str(r["Producto"])
        base_candidate = fallback_product_base(raw)
        if normalize(raw) in exclusiones or base_candidate in exclusiones:
            continue
        info = alias_lookup(raw, alias_map)
        base = info["Producto Base"]
        if base in exclusiones:
            continue
        qty_raw = safe_num(r["Total_Cantidad"], 0)
        qty, regla = convert_quantity(raw, base, qty_raw, info.get("Tipo Presentación", ""), info.get("Pack Detectado", np.nan), productos_map)
        rows.append({
            "Origen": "Ventas data", "Grupo": r.get("Grupo", ""), "SubRubro": r.get("SubCategoria", ""),
            "Producto Base": base, "Nombre original": raw, "Cantidad Eq": qty,
            "Regla Conversión": regla, "Alias Encontrado": info.get("Alias Encontrado", False)
        })
    return pd.DataFrame(rows)


def format_workbook(path):
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 14), 70)
        ws.freeze_panes = "A2"

    # Código Compra ancho fijo carrito Mark VIII
    if "Carrito" in wb.sheetnames:
        ws_carrito = wb["Carrito"]
        for row in ws_carrito.iter_rows():
            for cell in row:
                if str(cell.value).strip() == "Código Compra":
                    ws_carrito.column_dimensions[cell.column_letter].width = 15

    wb.save(path)



def procesar_costo_stock(stock_file, maestro_file, output_file, carrito_file=DEFAULT_CARRITO_FILE):
    """
    Genera un informe de valorización de stock usando solo:
    - archivo stock
    - maestro
    - archivo de carrito/precios
    """
    config = load_config(maestro_file)
    productos = load_productos(maestro_file)
    aliases = load_aliases(maestro_file)
    exclusiones = load_exclusiones(maestro_file)
    precios_carrito = load_precios_carrito(carrito_file)

    productos_validos = productos[(productos["Activo"]) & (~productos["Excluir"])].copy()
    productos_map = productos_validos.set_index("Producto Base").to_dict(orient="index")

    alias_map = {}
    for _, r in aliases.iterrows():
        alias_map[r["Alias Normalizado"]] = {
            "Producto Base": r["Producto Base"],
            "Tipo Presentación": r.get("Tipo Presentación", ""),
            "Pack Detectado": r.get("Pack Detectado", np.nan),
            "Regla Conversión": r.get("Regla Conversión", ""),
            "Alias Encontrado": True,
        }

    stock_rows = read_stock(stock_file, config, alias_map, productos_map, exclusiones)

    stock_grouped = stock_rows.groupby(["Grupo", "SubRubro", "Producto Base"], as_index=False).agg({
        "Cantidad Eq": "sum",
        "Nombre original": lambda x: " + ".join(sorted(set(x.astype(str))))
    }).rename(columns={
        "Cantidad Eq": "Stock",
        "Nombre original": "Productos agrupados"
    })

    final = stock_grouped.merge(productos_validos, on="Producto Base", how="left")

    final["Compra Mínima"] = pd.to_numeric(final["Compra Mínima"], errors="coerce").fillna(1)
    final["Código Compra Normalizado"] = final["Código Compra"].apply(normalizar_codigo_compra)
    final["Precio Pack"] = final["Código Compra Normalizado"].map(precios_carrito).fillna(0).astype(int)

    final["Costo Unitario Eq"] = final.apply(
        lambda r: r["Precio Pack"] / r["Compra Mínima"] if r["Compra Mínima"] else 0,
        axis=1
    )
    final["Valor Stock Actual"] = (final["Stock"] * final["Costo Unitario Eq"]).fillna(0).astype(int)

    detalle = final[[
        "Grupo",
        "SubRubro",
        "Producto Base",
        "Stock",
        "Compra Mínima",
        "Precio Pack",
        "Costo Unitario Eq",
        "Valor Stock Actual",
        "Código Compra",
        "Producto Compra",
        "Productos agrupados",
    ]].copy().rename(columns={
        "Producto Base": "Producto"
    }).sort_values("Valor Stock Actual", ascending=False)

    categoria = (
        detalle.groupby("Grupo", as_index=False)
        .agg({
            "Stock": "sum",
            "Valor Stock Actual": "sum",
        })
        .sort_values("Valor Stock Actual", ascending=False)
    )

    stock_negativo = detalle[pd.to_numeric(detalle["Stock"], errors="coerce").fillna(0) < 0].copy()

    resumen = pd.DataFrame({
        "Concepto": [
            "Valor stock actual",
            "Productos valorizados",
            "Productos con stock negativo",
        ],
        "Valor": [
            int(detalle["Valor Stock Actual"].sum()),
            len(detalle),
            len(stock_negativo),
        ]
    })

    explicacion = pd.DataFrame({
        "Explicación": [
            "INFORME COSTO STOCK",
            "Este informe valoriza el stock actual sin requerir archivos de ventas.",
            "Usa el archivo de stock, el maestro de productos y el archivo de carrito/precios.",
            "Precio pack se toma del carrito/precios por código de compra.",
            "Valor stock actual = Stock × (Precio pack / Compra mínima).",
            "Si hay productos con stock negativo, deben revisarse porque pueden afectar la valorización.",
        ]
    })

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        resumen.to_excel(writer, index=False, sheet_name="Resumen")
        categoria.to_excel(writer, index=False, sheet_name="Valorización categoría")
        detalle.to_excel(writer, index=False, sheet_name="Valorización producto")
        stock_negativo.to_excel(writer, index=False, sheet_name="Stock negativo")
        explicacion.to_excel(writer, index=False, sheet_name="Explicación")

    format_workbook(output_file)

    return {
        "detalle": detalle,
        "categoria": categoria,
        "stock_negativo": stock_negativo,
        "valor_stock_total": int(detalle["Valor Stock Actual"].sum()),
        "output_file": output_file,
    }


def procesar_archivos(stock_file, sabores_file, data_file, maestro_file, output_file, overrides=None, carrito_file=DEFAULT_CARRITO_FILE):
    config = load_config(maestro_file, overrides=overrides)
    productos = load_productos(maestro_file)
    aliases = load_aliases(maestro_file)
    exclusiones = load_exclusiones(maestro_file)
    precios_carrito = load_precios_carrito(carrito_file)

    productos_validos = productos[(productos["Activo"]) & (~productos["Excluir"])].copy()
    productos_map = productos_validos.set_index("Producto Base").to_dict(orient="index")

    alias_map = {}
    for _, r in aliases.iterrows():
        alias_map[r["Alias Normalizado"]] = {
            "Producto Base": r["Producto Base"],
            "Tipo Presentación": r.get("Tipo Presentación", ""),
            "Pack Detectado": r.get("Pack Detectado", np.nan),
            "Regla Conversión": r.get("Regla Conversión", ""),
            "Alias Encontrado": True,
        }

    stock_rows = read_stock(stock_file, config, alias_map, productos_map, exclusiones)
    sabores_rows = read_sabores(sabores_file, alias_map, productos_map, exclusiones)
    data_rows = read_data(data_file, config, alias_map, productos_map, exclusiones)

    stock_grouped = stock_rows.groupby(["Grupo", "SubRubro", "Producto Base"], as_index=False).agg({
        "Cantidad Eq": "sum",
        "Nombre original": lambda x: " + ".join(sorted(set(x.astype(str))))
    }).rename(columns={"Cantidad Eq": "Stock", "Nombre original": "Productos agrupados"})

    ventas_all = pd.concat([sabores_rows, data_rows], ignore_index=True)
    ventas_grouped = ventas_all.groupby("Producto Base", as_index=False)["Cantidad Eq"].sum().rename(columns={"Cantidad Eq": "Ventas período"})

    final = stock_grouped.merge(ventas_grouped, on="Producto Base", how="outer")
    final["Stock"] = final["Stock"].fillna(0)
    final["Grupo"] = final["Grupo"].fillna("")
    final["SubRubro"] = final["SubRubro"].fillna("")
    final["Productos agrupados"] = final["Productos agrupados"].fillna("Sin stock en archivo")
    final["Ventas período"] = final["Ventas período"].fillna(0)

    dias_analizados = safe_num(config.get("Días analizados", 14), 14)
    semanas_objetivo = safe_num(config.get("Semanas objetivo", 4), 4)
    tiempo_reposicion = safe_num(config.get("Tiempo de reposición", 1), 1)
    valor_sin_rotacion = safe_num(config.get("Valor sin rotación", 100), 100)

    objetivo_total = semanas_objetivo + tiempo_reposicion
    final["Venta Semanal"] = final["Ventas período"] / dias_analizados * 7

    def calc_semanas(row):
        if row["Venta Semanal"] == 0 and row["Stock"] > 0:
            return valor_sin_rotacion
        if row["Venta Semanal"] == 0:
            return np.nan
        return row["Stock"] / row["Venta Semanal"]

    final["Semanas Stock"] = final.apply(calc_semanas, axis=1)
    final["Reposición necesaria"] = np.maximum(0, (final["Venta Semanal"] * objetivo_total) - final["Stock"])

    final = final.merge(productos_validos, on="Producto Base", how="left")

    # Posibles faltantes:
    # En Grido, ventas = 0 rara vez significa demanda cero. Si además el stock es bajo,
    # lo marcamos para revisión manual sin modificar automáticamente la compra sugerida.
    def es_helado_granel(row):
        grupo = normalize(row.get("Grupo", ""))
        subrubro = normalize(row.get("SubRubro", ""))
        tipo = normalize(row.get("Tipo Producto", ""))
        producto = normalize(row.get("Producto Base", ""))
        return (
            "helado" in grupo
            or "sabores" in subrubro
            or "sabor" in tipo
            or "7 800" in producto
        )

    def observacion_faltante(row):
        ventas = safe_num(row.get("Ventas período", 0), 0)
        stock_actual = safe_num(row.get("Stock", 0), 0)
        if ventas != 0:
            return ""

        if es_helado_granel(row) and stock_actual < 1:
            return "Posible faltante: sin ventas y stock bajo"

        if (not es_helado_granel(row)) and stock_actual < 5:
            return "Posible faltante: sin ventas y stock bajo"

        return ""

    final["Observación"] = final.apply(observacion_faltante, axis=1)
    final["Compra Mínima"] = pd.to_numeric(final["Compra Mínima"], errors="coerce").fillna(1)

    final["Código Compra Normalizado"] = final["Código Compra"].apply(normalizar_codigo_compra)
    final["Precio Pack"] = final["Código Compra Normalizado"].map(precios_carrito).fillna(0)

    final["Packs a Comprar"] = final.apply(lambda r: ceil_to_int(r["Reposición necesaria"] / r["Compra Mínima"]) if r["Reposición necesaria"] > 0 else 0, axis=1)
    final["Unidades Finales"] = final["Packs a Comprar"] * final["Compra Mínima"]

    final["Costo Unitario Eq"] = final.apply(
        lambda r: r["Precio Pack"] / r["Compra Mínima"] if r["Compra Mínima"] else 0,
        axis=1
    )
    final["Valor Stock Actual"] = (final["Stock"] * final["Costo Unitario Eq"]).fillna(0).astype(int)
    final["Valor Pedido Sugerido"] = (final["Packs a Comprar"] * final["Precio Pack"]).fillna(0).astype(int)
    final["Precio Pack"] = final["Precio Pack"].fillna(0).astype(int)

    all_detected = pd.concat([
        stock_rows[["Origen", "Grupo", "SubRubro", "Producto Base", "Nombre original", "Regla Conversión", "Alias Encontrado"]],
        sabores_rows[["Origen", "Grupo", "SubRubro", "Producto Base", "Nombre original", "Regla Conversión", "Alias Encontrado"]],
        data_rows[["Origen", "Grupo", "SubRubro", "Producto Base", "Nombre original", "Regla Conversión", "Alias Encontrado"]],
    ], ignore_index=True)

    maestro_set = set(productos_validos["Producto Base"].dropna().astype(str))
    sin_clasificar = all_detected[~all_detected["Producto Base"].isin(maestro_set)].drop_duplicates().copy()

    pedido = final[[
        "Grupo", "Producto Base", "Stock", "Ventas período", "Venta Semanal", "Semanas Stock",
        "Reposición necesaria", "Compra Mínima", "Packs a Comprar", "Unidades Finales",
        "Precio Pack", "Valor Stock Actual", "Valor Pedido Sugerido",
        "Código Compra", "Producto Compra", "Observación", "Productos agrupados"
    ]].copy()

    pedido = pedido.rename(columns={
        "Producto Base": "Producto",
        "Ventas período": f"Ventas {int(dias_analizados)} días"
    })

    for col in [
        "Stock", f"Ventas {int(dias_analizados)} días", "Venta Semanal", "Semanas Stock",
        "Reposición necesaria", "Compra Mínima", "Unidades Finales",
        "Precio Pack", "Valor Stock Actual", "Valor Pedido Sugerido"
    ]:
        pedido[col] = pd.to_numeric(pedido[col], errors="coerce").round(2)

    pedido = pedido.sort_values(["Grupo", "Producto"])

    posibles_faltantes = pedido[
        pedido["Observación"].astype(str).str.contains("Posible faltante", na=False)
    ].copy()

    stock_negativo = pedido[
        pd.to_numeric(pedido["Stock"], errors="coerce").fillna(0) < 0
    ].copy()


    carrito = pedido[pedido["Packs a Comprar"] > 0][[
        "Código Compra",
        "Producto Compra",
        "Packs a Comprar",
        "Precio Pack",
        "Valor Pedido Sugerido",
        "Observación"
    ]].copy()

    valorizacion_categoria = (
        pedido.groupby("Grupo", as_index=False)
        .agg({
            "Valor Stock Actual": "sum",
            "Valor Pedido Sugerido": "sum",
            "Stock": "sum",
            "Packs a Comprar": "sum",
        })
        .sort_values("Valor Stock Actual", ascending=False)
    )

    valorizacion_producto = pedido[[
        "Grupo",
        "Producto",
        "Stock",
        "Precio Pack",
        "Valor Stock Actual",
        "Packs a Comprar",
        "Valor Pedido Sugerido",
        "Código Compra",
        "Producto Compra",
    ]].copy().sort_values("Valor Stock Actual", ascending=False)

    advertencias = []

    if len(stock_negativo) > 0:
        advertencias.append(f"⚠️ STOCK NEGATIVO DETECTADO: {len(stock_negativo)} productos. Revisar hoja 'Stock negativo'.")

    if len(sin_clasificar) > 0:
        advertencias.append(f"⚠️ PRODUCTOS SIN CLASIFICAR: {len(sin_clasificar)} productos. Revisar hoja 'Sin clasificar'.")

    if len(posibles_faltantes) > 0:
        advertencias.append(f"⚠️ POSIBLES FALTANTES: {len(posibles_faltantes)} productos. Revisar hoja 'Posibles faltantes'.")

    if not advertencias:
        advertencias = ["✅ No se detectaron alertas importantes."]

    advertencias_df = pd.DataFrame({"Carrito": advertencias})

    explicacion = pd.DataFrame({
        "Explicación": [
            "ARCHIVOS DE ENTRADA",
            "stock.csv: stock actual del sistema. Se suma Stock + Tránsito del archivo de stock.",
            "cajas_por_sabor.xlsx: ventas de sabores 7,800 kg.",
            "data.xlsx: exportación Power BI con impulsivos y congelados.",
            "Maestro_Productos_Grido.xlsx: productos, códigos, compra mínima, aliases, exclusiones y configuración.",
            "",
            "CONFIGURACIÓN USADA",
            f"Semanas objetivo: {semanas_objetivo}",
            f"Tiempo de reposición: {tiempo_reposicion}",
            f"Días analizados: {dias_analizados}",
            f"Valor sin rotación: {valor_sin_rotacion}",
            "",
            "FÓRMULAS",
            "Venta semanal = Ventas del período / Días analizados × 7",
            "Objetivo total = Semanas objetivo + Tiempo de reposición",
            "Reposición necesaria = (Venta semanal × Objetivo total) - Stock",
            "Packs a comprar = redondeo hacia arriba(Reposición necesaria / Compra mínima)",
            "Unidades finales = Packs a comprar × Compra mínima",
            "Valor pedido sugerido = Packs a comprar × Precio pack",
            "Valor stock actual = Stock × (Precio pack / Compra mínima)",
            "",
            "ACLARACIÓN",
            "Tiempo de reposición NO es la columna Tránsito del archivo stock.",
            "La columna Tránsito del stock representa mercadería ya pedida/en camino y se suma al stock.",
            "",
            "PRODUCTOS NUEVOS",
            "Si aparece un producto no encontrado en el maestro, se informa en la hoja Sin clasificar.",
            "No se debería comprar automáticamente un producto sin clasificar hasta cargarlo en el maestro.",
            "",
            "POSIBLES FALTANTES",
            "Si un producto no tuvo ventas y además tiene stock bajo, se marca como posible faltante.",
            "Esta alerta no modifica automáticamente la compra sugerida. Sirve para revisión manual.",
            "Para helado a granel se considera stock bajo si es menor a 1.",
            "Para el resto de productos se considera stock bajo si es menor a 5.",
        ]
    })

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        advertencias_df.to_excel(writer, index=False, sheet_name="Carrito", startrow=0)
        carrito.to_excel(writer, index=False, sheet_name="Carrito", startrow=len(advertencias_df) + 3)

        valorizacion_categoria.to_excel(writer, index=False, sheet_name="Valorización categoría")
        valorizacion_producto.to_excel(writer, index=False, sheet_name="Valorización producto")

        pedido.to_excel(writer, index=False, sheet_name="Pedido Final")
        stock_negativo.to_excel(writer, index=False, sheet_name="Stock negativo")
        posibles_faltantes.to_excel(writer, index=False, sheet_name="Posibles faltantes")
        sin_clasificar.to_excel(writer, index=False, sheet_name="Sin clasificar")
        explicacion.to_excel(writer, index=False, sheet_name="Explicación")

    format_workbook(output_file)

    return {
        "pedido": pedido,
        "stock_negativo": stock_negativo,
        "posibles_faltantes": posibles_faltantes,
        "sin_clasificar": sin_clasificar,
        "output_file": output_file,
        "config": config,
        "valor_stock_total": float(pd.to_numeric(pedido["Valor Stock Actual"], errors="coerce").fillna(0).sum()),
        "valor_pedido_total": float(pd.to_numeric(pedido["Valor Pedido Sugerido"], errors="coerce").fillna(0).sum()),
        "valorizacion_categoria": valorizacion_categoria,
        "valorizacion_producto": valorizacion_producto,
    }
